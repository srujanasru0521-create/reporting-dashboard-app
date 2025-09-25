import csv
from openpyxl import Workbook
from fpdf import FPDF

def generate_csv(data, file_path):
    """Generates a CSV report."""
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        if not data:
            csvfile.write("No data available for the selected criteria\n")
            return
        fieldnames = data[0].keys()
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)

def generate_xlsx(data, file_path):
    """Generates an XLSX report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"
    
    if not data:
        ws.append(["No data available for the selected criteria"])
    else:
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append(list(row.values()))
    
    wb.save(file_path)

def generate_pdf(data, file_path):
    """Generates a PDF report."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    if not data:
        pdf.cell(200, 10, txt="No data available for the selected criteria.", ln=True, align='C')
    else:
        # Add title
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(200, 10, txt="Financial Report", ln=True, align='C')
        pdf.ln(10)
        
        # Reset font for content
        pdf.set_font("Arial", size=10)
        
        # Get headers and calculate column width
        headers = list(data[0].keys())
        col_width = 190 / len(headers)  # Distribute width across page
        
        # Add headers
        pdf.set_font("Arial", 'B', 10)
        for header in headers:
            pdf.cell(col_width, 10, txt=str(header)[:15], border=1, align='C')
        pdf.ln()
        
        # Add data rows
        pdf.set_font("Arial", size=9)
        for row in data:
            for value in row.values():
                # Truncate long values to fit in cells
                display_value = str(value)[:15] if value is not None else ""
                pdf.cell(col_width, 10, txt=display_value, border=1, align='C')
            pdf.ln()
    
    pdf.output(file_path)